import os
from openpyxl import Workbook
from openpyxl import load_workbook

def marksheet():
    out_folder =  "output_folder"
    if not os.path.exists(out_folder):
        os.mkdir(out_folder)
    
    grade_dict = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,
            'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}
    rollfile = open ('names-roll.csv', 'r')
    rolls_f = rollfile.readlines()
    rolls = [i[:-1].split(",") for i in rolls_f]
    gradesfile = open ('grades.csv', 'r')
    grades_f = gradesfile.readlines()
    grades = [i[:-1].split(",") for i in grades_f]
    subfile = open( 'subjects_master.csv','r')
    subs_f = subfile.readlines()
    subs = [i[:-1].split(",") for i in subs_f]
    # print(subs[:10])
    # print(rolls[:10])
    # print(grades[:10])
    sub_dic = {}
    for c in subs:
        if c[0] not in sub_dic:
            sub_dic[c[0]] = c
    # head part in the marksheet
    head= ["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"]
    for r in rolls[1:]:
        i=1
        wb= Workbook()
        while i<=8 :
        	sem = 'Sem'+str(i)
        	if sem not in wb.sheetnames:
        		wb.create_sheet(sem,i)
        	ws = wb[sem]
        	ws.append(head)
        	j=1
        	for x in grades[1:]:
        		if x[0]==r[0] and int(x[1])==i:
        			d=list()
        			d.append(j)
        			d.append(x[2])
        			y=sub_dic[x[2]]
        			order = [1,2,3]
        			for k in order:
        				d.append(y[k])
        			d.append(x[5])
        			d.append(x[4])
        			ws.append(d)
        			j+=1
        	i+=1
        	wb.save(f'output_folder\\{r[0]}.xlsx')
    return
marksheet()