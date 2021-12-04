import csv,os,shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
if  os.path.exists("output"): shutil.rmtree("output")

os.makedirs("output")
with open("names-roll.csv", 'r') as file:
    rows = csv.reader(file)
    rlnm = {line[0]: line[1].strip() for line in rows if line[0] != "Roll"}
with open("subjects_master.csv", 'r') as file:
    rows = csv.reader(file)
    sbj = {line[0]: [line[1], line[2]] for line in rows if line[0] != "subno"}
rplc = {"AA": "10", "AB": "9", "BB": "8", "BC": "7",
        "CC": "6", "CD": "5", "DD": "4", "F": "0", "I": "0"}
s = {i: [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] for i in rlnm}
u = {i: [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] for i in rlnm}
with open("grades.csv", 'r') as file:
    rows = csv.reader(file)
    for line in rows:
        if line[0] == 'Roll': continue
        s[line[0]][int(line[1])] = s[line[0]][int(line[1])]+int(line[3])
        u[line[0]][int(line[1])] = u[line[0]][int(line[1])] +int(line[3])*int(rplc[line[4].strip().strip("*")])
        with open(f"output\\{line[0]}_{line[1]}.csv", "a", newline='') as file:
            writer = csv.writer(file)
            writer.writerow([line[2], sbj[line[2]][0],sbj[line[2]][1], line[3], line[5], line[4]])
r = {key: [i for i, el in enumerate(s[key]) if el != 0] for key in rlnm}
for key in rlnm:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    for x in r[key]:
        wb.create_sheet(index=x, title=f"Sem{x}")
        sheet = wb[f"Sem{x}"]
        sheet.append(["Sl No.", "Subject No.", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade"])
        with open(f"output\\{key}_{x}.csv", 'r') as file:
            rows = csv.reader(file)
            for line in rows:
                row_count = sheet.max_row
                line.insert(0, row_count)
                sheet.append(line)
        os.remove(f"output\\{key}_{x}.csv")
    sheet = wb["Sem2"]
    row_count = sheet.max_row
    for row in sheet.iter_rows(min_row=1, max_row=row_count):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor="B5DDFB", fill_type="solid")
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill(fgColor="77C3FD", fill_type="solid")
    wb.create_sheet(index=0, title="Overall")
    Overall = wb["Overall"]
    Overall.append(["Roll No.", key])
    Overall.append(["Name of Student", rlnm[key]])
    Overall.append(["Discipline", key[4:6]])
    Overall.append(["Semester No."]+[i for i in r[key]])
    Overall.append(["Semester wise Credit Taken"]+[s[key][i] for i in r[key]])
    Overall.append(["SPI"]+[round(u[key][i]/s[key][i], 2) for i in r[key]])
    Overall.append(["Total Credits Taken"]+[sum(s[key][:i+1]) for i in r[key]])
    Overall.append(["CPI"]+[round(sum(u[key][:i+1])/sum(s[key][:i+1]), 2) for i in r[key]])
    wb.save(f"output//{key}.xlsx")