'''  Team Members
PCS ASWIN KUMAR - 1901MM25
BADRI SAI PRASAD - 1901MM11
'''

from re import L
from flask import Flask, render_template, request, redirect, url_for,flash

from openpyxl import Workbook
import os
import csv
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import shutil
from reportlab.lib import colors
from reportlab.lib.pagesizes import A1, A2, A3, A4, A5
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from datetime import datetime
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from openpyxl import Workbook
import os
import csv
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from fpdf import FPDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A1, A2, A3, A4, A5
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from datetime import datetime
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
path = os.getcwd()
range1 = '0401ME10'
range2 = '0401ME15'
roll1 = 1
roll2 = 6
l_range = []
def generate_marksheet():

 if  not os.path.exists("Transcripts-IITP") : os.makedirs("output")
 with open("names-roll.csv", 'r') as file:
    rows = csv.reader(file)
    rlnm = {line[0]: line[1].strip() for line in rows if line[0] != "Roll"}
 with open("subjects_master.csv", 'r') as file:
    rows = csv.reader(file)
    sbj = {line[0]: [line[1], line[2]] for line in rows if line[0] != "subno"}
 rplc = {"AA": "10", "AB": "9", "BB": "8", "BC": "7",
        "CC": "6", "CD": "5", "DD": "4", "F": "0", "I": "0"}
 s = {i: [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] for i in rlnm}
 u = {i: [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] for i in rlnm}
 with open("grades.csv", 'r') as file:
    rows = csv.reader(file)
    for line in rows:
        if line[0] == 'Roll': continue
        s[line[0]][int(line[1])] = s[line[0]][int(line[1])]+int(line[3])
        u[line[0]][int(line[1])] = u[line[0]][int(line[1])] +int(line[3])*int(rplc[line[4].strip().strip("*")])
        with open(f"output\\{line[0]}_{line[1]}.csv", "a", newline='') as file:
            writer = csv.writer(file)
            writer.writerow([line[2], sbj[line[2]][0],sbj[line[2]][1], line[3], line[5], line[4]])
 r = {key: [i for i, el in enumerate(s[key]) if el != 0] for key in rlnm}
 for key in rlnm:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    for x in r[key]:
        wb.create_sheet(index=x, title=f"Sem{x}")
        sheet = wb[f"Sem{x}"]
        sheet.append(["Sl No.", "Subject No.", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade"])
        with open(f"output\\{key}_{x}.csv", 'r') as file:
            rows = csv.reader(file)
            for line in rows:
                row_count = sheet.max_row
                line.insert(0, row_count)
                sheet.append(line)
        os.remove(f"output\\{key}_{x}.csv")
    sheet = wb["Sem2"]
    row_count = sheet.max_row
    for row in sheet.iter_rows(min_row=1, max_row=row_count):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fgColor="B5DDFB", fill_type="solid")
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill(fgColor="77C3FD", fill_type="solid")
    wb.create_sheet(index=0, title="Overall")
    Overall = wb["Overall"]
    Overall.append(["Roll No.", key])
    Overall.append(["Name of Student", rlnm[key]])
    Overall.append(["Discipline", key[4:6]])
    Overall.append(["Semester No."]+[i for i in r[key]])
    Overall.append(["Semester wise Credit Taken"]+[s[key][i] for i in r[key]])
    Overall.append(["SPI"]+[round(u[key][i]/s[key][i], 2) for i in r[key]])
    Overall.append(["Total Credits Taken"]+[sum(s[key][:i+1]) for i in r[key]])
    Overall.append(["CPI"]+[round(sum(u[key][:i+1])/sum(s[key][:i+1]), 2) for i in r[key]])
    wb.save(f"output//{key}.xlsx")

                 
 return 





def generate_transcript() :
    global path            
    os.chdir(path)
    #generate_marksheet()
    if  os.path.exists("Transcripts-IITP"): shutil.rmtree("Transcripts-IITP")
    path = os.getcwd()
    #os.mkdir('Transcripts-IITP')
    src = os.getcwd() + '\output'
    dst = os.getcwd() + '\Transcripts-IITP'
    shutil.copytree(src, dst)
    os.chdir( dst)
    list = os.listdir()


    
    for l in list :
      c = canvas.Canvas(l[:8] + '.pdf', pagesize=A2)  
      width, height = A2  
      excel_file = openpyxl.load_workbook(l)
      overall1,overall2,overall3,overall4,overall5,overall6,overall7,overall8= [],[],[],[],[],[],[],[]
      if 'Sem1' in excel_file.sheetnames :
       Sem1 = excel_file['Sem1']
       for row in Sem1.iter_rows(min_row=Sem1.min_row, min_col=Sem1.min_column, max_row=Sem1.max_row, max_col=Sem1.max_column):
         sem1= []
         for cell in row:
             cell.font = Font(size=20)
             sem1.append(cell.value)
        
         overall1.append(sem1)
       for j in overall1 :
          del j[0]  
          del j[4] 
       t1=Table(overall1, rowHeights=15)
       t1.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.2, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t1.wrapOn(c, width, height)
       t1.drawOn(c, 20*mm, 475*mm)
       t1._elementWidth
         
      
      if 'Sem2' in excel_file.sheetnames :
       Sem2 = excel_file['Sem2']
       for row in Sem2.iter_rows(min_row=Sem2.min_row, min_col=Sem2.min_column, max_row=Sem2.max_row, max_col=Sem2.max_column):
         sem2= []
         for cell in row:
             
             sem2.append(cell.value)
         
         overall2.append(sem2)
       for j in overall2:
          del j[0]  
          del j[4] 
       t2=Table(overall2, rowHeights=15)
       t2.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t2.wrapOn(c, width, height)
       t2.drawOn(c, 150*mm, 475*mm)
       
      
      if 'Sem3' in excel_file.sheetnames :
       Sem3 = excel_file['Sem3']
       for row in Sem3.iter_rows(min_row=Sem3.min_row, min_col=Sem3.min_column, max_row=Sem3.max_row, max_col=Sem3.max_column):
         sem3= []
         for cell in row:
             
             sem3.append(cell.value)
         
         overall3.append(sem3)
       for j in overall3 :
          del j[0]  
          del j[4]
       t3=Table(overall3, rowHeights=15)
       t3.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t3.wrapOn(c, width, height)
       t3.drawOn(c, 260*mm, 475*mm)
        
      
      if 'Sem4' in excel_file.sheetnames :
       Sem4 = excel_file['Sem4']
       for row in Sem4.iter_rows(min_row=Sem4.min_row, min_col=Sem4.min_column, max_row=Sem4.max_row, max_col=Sem4.max_column):
         sem4= []
         for cell in row:
             
             sem4.append(cell.value)
         
         overall4.append(sem4)
       for j in overall4 :
          del j[0]  
          del j[4]
       t4=Table(overall4, rowHeights=15)
       t4.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t4.wrapOn(c, width, height)
       t4.drawOn(c, 20*mm, 395*mm)
       

      if 'Sem5' in excel_file.sheetnames :
       Sem5 = excel_file['Sem5']
       for row in Sem5.iter_rows(min_row=Sem5.min_row, min_col=Sem5.min_column, max_row=Sem5.max_row, max_col=Sem5.max_column):
         sem5= []
         for cell in row:
             
             sem5.append(cell.value)
         
         overall5.append(sem5)
       for j in overall5 :
          del j[0]  
          del j[4]
       t5=Table(overall5, rowHeights=15,)
       t5.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t5.wrapOn(c, width, height)
       t5.drawOn(c, 160*mm, 400*mm)
         
      
      if 'Sem6' in excel_file.sheetnames :
       Sem6 = excel_file['Sem6']
       for row in Sem6.iter_rows(min_row=Sem6.min_row, min_col=Sem6.min_column, max_row=Sem6.max_row, max_col=Sem6.max_column):
         sem6= []
         for cell in row:
             
             sem6.append(cell.value)
         
         overall6.append(sem6)
       for j in overall6 :
          del j[0]  
          del j[4]
       t6=Table(overall6, rowHeights=15)
       t6.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t6.wrapOn(c, width, height)
       t6.drawOn(c, 280*mm, 400*mm)
          

      if 'Sem7' in excel_file.sheetnames :
       Sem7 = excel_file['Sem7']
       for row in Sem7.iter_rows(min_row=Sem7.min_row, min_col=Sem7.min_column, max_row=Sem7.max_row, max_col=Sem7.max_column):
         sem7= []
         for cell in row:
             
             sem7.append(cell.value)
         
         overall7.append(sem7)
       for j in overall7 :
          del j[0]  
          del j[4] 
       t7=Table(overall7, rowHeights=15)
       t7.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t7.wrapOn(c, width, height)
       t7.drawOn(c, 20*mm, 320*mm)
         
     
      if 'Sem8' in excel_file.sheetnames :
        Sem8 = excel_file['Sem8']
        for row in Sem8.iter_rows(min_row=Sem8.min_row, min_col=Sem8.min_column, max_row=Sem8.max_row, max_col=Sem8.max_column):
          sem8= []
          for cell in row:
             
             sem8.append(cell.value)
         
          overall8.append(sem8)
        for j in overall8 :
          del j[0]  
          del j[4] 
        t8=Table(overall8 , rowHeights=15)
        t8.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
        t8.wrapOn(c, width, height)
        t8.drawOn(c, 150*mm, 320*mm)
          

      credits_taken = excel_file['Overall']
      sem1_credits   = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 2).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 2).value) + ' cpi :'+ str(credits_taken.cell(row = 8 , column = 2).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 2).value)
      sem2_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 3).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 3).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 3).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 3).value)
      sem3_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 4).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 4).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 4).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 4).value)
      sem4_credits = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 5).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 5).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 5).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 5).value)
      sem5_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 6).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 6).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 6).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 6).value)
      sem6_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 7).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 7).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 7).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 7).value)
      sem7_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 8).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 8).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 8).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 8).value)
      sem8_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 9).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 9).value) + ' cpi :'+ str(credits_taken.cell(row = 8 , column = 9).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 9).value)
      ''' crd1,crd2,crd3,crd4,crd5,crd6,crd7,crd8, = [],[],[],[],[],[],[],[]
      crd1.append(sem1_credits),crd1.append(creidts_cleared1),crd1.append(spi1),crd1.append(cpi1)
      crd2.append(sem2_credits),crd1.append(creidts_cleared2),crd1.append(spi2),crd1.append(cpi2)
      crd3.append(sem3_credits),crd1.append(creidts_cleared3),crd1.append(spi3),crd1.append(cpi3)
      crd4.append(sem4_credits),crd1.append(creidts_cleared4),crd1.append(spi4),crd1.append(cpi4)
      crd5.append(sem5_credits),crd1.append(creidts_cleared5),crd1.append(spi5),crd1.append(cpi5)
      crd6.append(sem6_credits),crd1.append(creidts_cleared6),crd1.append(spi6),crd1.append(cpi6)
      crd7.append(sem7_credits),crd1.append(creidts_cleared7),crd1.append(spi7),crd1.append(cpi7)
      crd8.append(sem8_credits),crd1.append(creidts_cleared8),crd1.append(spi8),crd1.append(cpi8)
      print(crd1)'''
      string ='     roll no : ' + str(credits_taken.cell(row = 1 , column = 2).value) + '                                       Name: ' + str(credits_taken.cell(row = 2 , column = 2).value) + '                  Year of admission :' + '20' + str(credits_taken.cell(row = 1 , column = 2).value[0]) + str(credits_taken.cell(row = 1 , column = 2).value[1])
      programme = '     programme : ' + 'Bachelor of Technology' + '                 Discpline:' + str(credits_taken.cell(row = 3 , column = 2).value)


      

      

     
      #c.setFontSize(10)
      #c.setFont("Times-Roman", 18)
      c.rect(30,40,1100,1600) # border
      c.rect(30,1558,1100,0)#line below image
      c.rect(30,1080,1100,0) # line above 7 and 8 semester
      c.rect(250,1500,670,40) #table below image
      c.drawString(250,1520,string)
      c.drawString(250,1505,programme)
      c.rect(30,1300,1100,0) # line above 4,5 and 6 semester
      c.rect(30,850,1100,0) # line below 7 and 8 semester
      c.drawString(40,700,'Date Generated :' + datetime.now().strftime("%d %b %Y, %H:%M"))
      c.drawString(800,700,'Assistant Registrar(Academic)')
      c.rect(59,1320,300,20) # for semester 1
      c.drawString(60,1485,'semester1')
      c.drawString(60,1325,sem1_credits)
     # c.drawBoundary(sem1_credits ,59,1320,300,20,)
      if 'Sem2' in excel_file.sheetnames :
       c.rect(430,1320,300,20) # for semester 2
       c.drawString(430,1325,sem2_credits)
       c.drawString(430,1485,'semester2')
      if 'Sem3' in excel_file.sheetnames :
       c.rect(760,1320,300,20) # for semester 3
       c.drawString(760,1325,sem3_credits)
       c.drawString(760,1485,'semester3')
      if 'Sem4' in excel_file.sheetnames :
       c.rect(59,1100,300,20) # for semester 4
       c.drawString(60,1105,sem4_credits)
       c.drawString(60,1280,'semester4')
      if 'Sem5' in excel_file.sheetnames :
       c.rect(430,1100,300,20) # for semester 5
       c.drawString(430,1105,sem5_credits)
       c.drawString(430,1280,'semester5')
      if 'Sem6' in excel_file.sheetnames :
       c.rect(760,1100,300,20) # for semester 6
       c.drawString(760,1105,sem6_credits)
       c.drawString(760,1280,'semester6')
      if 'Sem7' in excel_file.sheetnames :
       c.rect(59,880,300,20) # for semester 7
       c.drawString(60,885,sem7_credits)
       c.drawString(60,1050,'semester7')
      if 'Sem8' in excel_file.sheetnames :
       c.rect(430,880,300,20) # for semester 8
       c.drawString(430,885,sem8_credits)
       c.drawString(430,1050,'semester8')
      
      os.chdir(path)
      c.drawImage('transcript_logo.png', 32, 1559,width = 1095)
      c.drawImage('seal.png',460,650)
      c.drawImage('signature.png',800,710)
      os.chdir(dst)
      
      
     

      
      
      
     
      
      
      
      c.save()
      os.remove(l)

     
     

    return

def generate_range_transcript() :
    
    global path            
    os.chdir(path)
    generate_marksheet()
    path = os.getcwd()
    #os.mkdir('Transcripts-IITP')
    if  os.path.exists("Transcripts-IITP"): shutil.rmtree("Transcripts-IITP")
    src = os.getcwd() + '\output'
    dst = os.getcwd() + '\Transcripts-IITP'
    shutil.copytree(src, dst)
    os.chdir( dst)
    list = os.listdir()
    a=[]
    global range1
    global range2
    if range1 + '.xlsx' not in list :
      range1 = range1[:7] + str(int(range1[-1])+1)
    index1 = list.index(range1 + '.xlsx')
    if range2 + '.xlsx' not in list :
      range2 = range2[:7] + str(int(range2[-1])-1)
    index2 = list.index(range2 + '.xlsx')
    global roll1
    global roll2
    string = range1[:6]
    roll1 = int(range1[-2:])
    roll2 = int(range2[-1:])
    roll2 = roll2 + 1
    roll1 = int(roll1)
    roll2 = int(roll2)
    
    a = []
    b=[]
    #for i in range(roll1,roll2) :
     # if i not in list[index1:index2+1] :
       # print(i)
    for l in list[:index1] :
     os.remove(l)
    for l in list[index2+1:] :
     os.remove(l)
    for l in list[index1:index2+1] :
      c = canvas.Canvas(l[:8] + '.pdf', pagesize=A2)  
      width, height = A2  
      excel_file = openpyxl.load_workbook(l)
      overall1,overall2,overall3,overall4,overall5,overall6,overall7,overall8= [],[],[],[],[],[],[],[]
      if 'Sem1' in excel_file.sheetnames :
       Sem1 = excel_file['Sem1']
       for row in Sem1.iter_rows(min_row=Sem1.min_row, min_col=Sem1.min_column, max_row=Sem1.max_row, max_col=Sem1.max_column):
         sem1= []
         for cell in row:
             cell.font = Font(size=20)
             sem1.append(cell.value)
         
         overall1.append(sem1)
       for j in overall1 :
          del j[0]  
          del j[4] 
       t1=Table(overall1, rowHeights=15)
       t1.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.2, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t1.wrapOn(c, width, height)
       t1.drawOn(c, 20*mm, 475*mm)
       t1._elementWidth
         
      
      if 'Sem2' in excel_file.sheetnames :
       Sem2 = excel_file['Sem2']
       for row in Sem2.iter_rows(min_row=Sem2.min_row, min_col=Sem2.min_column, max_row=Sem2.max_row, max_col=Sem2.max_column):
         sem2= []
         for cell in row:
             
             sem2.append(cell.value)
         
         overall2.append(sem2)
       for j in overall2:
          del j[0]  
          del j[4] 
       t2=Table(overall2, rowHeights=15)
       t2.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t2.wrapOn(c, width, height)
       t2.drawOn(c, 150*mm, 475*mm)
       
      
      if 'Sem3' in excel_file.sheetnames :
       Sem3 = excel_file['Sem3']
       for row in Sem3.iter_rows(min_row=Sem3.min_row, min_col=Sem3.min_column, max_row=Sem3.max_row, max_col=Sem3.max_column):
         sem3= []
         for cell in row:
             
             sem3.append(cell.value)
         
         overall3.append(sem3)
       for j in overall3 :
          del j[0]  
          del j[4]
       t3=Table(overall3, rowHeights=15)
       t3.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t3.wrapOn(c, width, height)
       t3.drawOn(c, 260*mm, 475*mm)
        
      
      if 'Sem4' in excel_file.sheetnames :
       Sem4 = excel_file['Sem4']
       for row in Sem4.iter_rows(min_row=Sem4.min_row, min_col=Sem4.min_column, max_row=Sem4.max_row, max_col=Sem4.max_column):
         sem4= []
         for cell in row:
             
             sem4.append(cell.value)
         
         overall4.append(sem4)
       for j in overall4 :
          del j[0]  
          del j[4]
       t4=Table(overall4, rowHeights=15)
       t4.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t4.wrapOn(c, width, height)
       t4.drawOn(c, 20*mm, 395*mm)
       
      
      if 'Sem5' in excel_file.sheetnames :
       Sem5 = excel_file['Sem5']
       for row in Sem5.iter_rows(min_row=Sem5.min_row, min_col=Sem5.min_column, max_row=Sem5.max_row, max_col=Sem5.max_column):
         sem5= []
         for cell in row:
             
             sem5.append(cell.value)
         
         overall5.append(sem5)
       for j in overall5 :
          del j[0]  
          del j[4]
       t5=Table(overall5, rowHeights=15,)
       t5.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t5.wrapOn(c, width, height)
       t5.drawOn(c, 160*mm, 405*mm)
         
      
      if 'Sem6' in excel_file.sheetnames :
       Sem6 = excel_file['Sem6']
       for row in Sem6.iter_rows(min_row=Sem6.min_row, min_col=Sem6.min_column, max_row=Sem6.max_row, max_col=Sem6.max_column):
         sem6= []
         for cell in row:
             
             sem6.append(cell.value)
         
         overall6.append(sem6)
       for j in overall6 :
          del j[0]  
          del j[4]
       t6=Table(overall6, rowHeights=15)
       t6.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t6.wrapOn(c, width, height)
       t6.drawOn(c, 280*mm, 405*mm)
          
      
      if 'Sem7' in excel_file.sheetnames :
       Sem7 = excel_file['Sem7']
       for row in Sem7.iter_rows(min_row=Sem7.min_row, min_col=Sem7.min_column, max_row=Sem7.max_row, max_col=Sem7.max_column):
         sem7= []
         for cell in row:
             
             sem7.append(cell.value)
         
         overall7.append(sem7)
       for j in overall7 :
          del j[0]  
          del j[4] 
       t7=Table(overall7, rowHeights=15)
       t7.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
       t7.wrapOn(c, width, height)
       t7.drawOn(c, 20*mm, 325*mm)
         
      
      if 'Sem8' in excel_file.sheetnames :
        Sem8 = excel_file['Sem8']
        for row in Sem8.iter_rows(min_row=Sem8.min_row, min_col=Sem8.min_column, max_row=Sem8.max_row, max_col=Sem8.max_column):
          sem8= []
          for cell in row:
             
             sem8.append(cell.value)
         
          overall8.append(sem8)
        for j in overall8 :
          del j[0]  
          del j[4] 
        t8=Table(overall8 , rowHeights=15)
        t8.setStyle(TableStyle([('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                       ('BOX', (0,0), (-1,-1), 0.25, colors.black),('FONTSIZE', (0, 0), (-1, -1), 8),
                       ]))
        t8.wrapOn(c, width, height)
        t8.drawOn(c, 150*mm, 334*mm)
          
      
      credits_taken = excel_file['Overall']
      sem1_credits   = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 2).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 2).value) + ' cpi :'+ str(credits_taken.cell(row = 8 , column = 2).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 2).value)
      sem2_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 3).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 3).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 3).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 3).value)
      sem3_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 4).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 4).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 4).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 4).value)
      sem4_credits = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 5).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 5).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 5).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 5).value)
      sem5_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 6).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 6).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 6).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 6).value)
      sem6_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 7).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 7).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 7).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 7).value)
      sem7_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 8).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 8).value)  +' cpi :'+ str(credits_taken.cell(row = 8 , column = 8).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 8).value)
      sem8_credits  = 'credits_taken :' + str(credits_taken.cell(row = 5 , column = 9).value) + ' spi :' + str(credits_taken.cell(row = 6 , column = 9).value) + ' cpi :'+ str(credits_taken.cell(row = 8 , column = 9).value)+' creidts_cleared :'+str(credits_taken.cell(row = 5 , column = 9).value)
      ''' crd1,crd2,crd3,crd4,crd5,crd6,crd7,crd8, = [],[],[],[],[],[],[],[]
      crd1.append(sem1_credits),crd1.append(creidts_cleared1),crd1.append(spi1),crd1.append(cpi1)
      crd2.append(sem2_credits),crd1.append(creidts_cleared2),crd1.append(spi2),crd1.append(cpi2)
      crd3.append(sem3_credits),crd1.append(creidts_cleared3),crd1.append(spi3),crd1.append(cpi3)
      crd4.append(sem4_credits),crd1.append(creidts_cleared4),crd1.append(spi4),crd1.append(cpi4)
      crd5.append(sem5_credits),crd1.append(creidts_cleared5),crd1.append(spi5),crd1.append(cpi5)
      crd6.append(sem6_credits),crd1.append(creidts_cleared6),crd1.append(spi6),crd1.append(cpi6)
      crd7.append(sem7_credits),crd1.append(creidts_cleared7),crd1.append(spi7),crd1.append(cpi7)
      crd8.append(sem8_credits),crd1.append(creidts_cleared8),crd1.append(spi8),crd1.append(cpi8)
      print(crd1)'''
      string ='     roll no : ' + str(credits_taken.cell(row = 1 , column = 2).value) + '                                       Name: ' + str(credits_taken.cell(row = 2 , column = 2).value) + '                  Year of admission :' + '20' + str(credits_taken.cell(row = 1 , column = 2).value[0]) + str(credits_taken.cell(row = 1 , column = 2).value[1])
      programme = '     programme : ' + 'Bachelor of Technology' + '                 Discpline:' + str(credits_taken.cell(row = 3 , column = 2).value)


      

      

     
      #c.setFontSize(10)
      #c.setFont("Times-Roman", 18)
      c.rect(30,40,1100,1600) # border
      c.rect(30,1558,1100,0)#line below image
      c.rect(30,1080,1100,0) # line above 7 and 8 semester
      c.rect(250,1500,670,40) #table below image
      c.drawString(250,1520,string)
      c.drawString(250,1505,programme)
      c.rect(30,1300,1100,0) # line above 4,5 and 6 semester
      c.rect(30,850,1100,0) # line below 7 and 8 semester
      c.drawString(40,700,'Date Generated :' + datetime.now().strftime("%d %b %Y, %H:%M"))
      c.drawString(800,700,'Assistant Registrar(Academic)')
      c.rect(59,1320,300,20) # for semester 1
      c.drawString(60,1485,'semester1')
      c.drawString(60,1325,sem1_credits)
     # c.drawBoundary(sem1_credits ,59,1320,300,20,)
      c.rect(430,1320,300,20) # for semester 2
      c.drawString(430,1325,sem2_credits)
      c.drawString(430,1485,'semester2')
      c.rect(760,1320,300,20) # for semester 3
      c.drawString(760,1325,sem3_credits)
      c.drawString(760,1485,'semester3')
      c.rect(59,1100,300,20) # for semester 4
      c.drawString(60,1105,sem4_credits)
      c.drawString(60,1280,'semester4')
      c.rect(430,1100,300,20) # for semester 5
      c.drawString(430,1105,sem5_credits)
      c.drawString(430,1280,'semester5')
      c.rect(760,1100,300,20) # for semester 6
      c.drawString(760,1105,sem6_credits)
      c.drawString(760,1280,'semester6')
      c.rect(59,880,300,20) # for semester 7
      c.drawString(60,885,sem7_credits)
      c.drawString(60,1050,'semester7')
      c.rect(430,880,300,20) # for semester 8
      c.drawString(430,885,sem8_credits)
      c.drawString(430,1050,'semester8')
      
      os.chdir(path)
      c.drawImage('transcript_logo.png', 32, 1559,width = 1095)
      c.drawImage('seal.png',430,650)
      c.drawImage('signature.png',800,710)
      os.chdir(dst)
      
      
     

      
      
      
     
      
      
      
      c.save()
      os.remove(l)

     
     


    return


app = Flask(__name__)
app.secret_key ='dont tell'

@app.route('/')
def index():
     return render_template('index.html')

@app.route('/', methods=['POST'])
def upload_file():
     uploaded_file = request.files['file']
     if uploaded_file.filename != '':
        uploaded_file.save(uploaded_file.filename)
     return redirect(url_for('index'))
def not_range() :
    global path
    os.chdir(path)
    global range1
    global range2
    from_range = range1
    to_range = range2
    from_RANGE = from_range.upper()
    to_RANGE = to_range.upper()
    h1 = int(to_RANGE[-2:])
    he = from_RANGE[:-2]
    hem = int(from_RANGE[-2:])
    print(hem, type(hem))
    print(he, type(he))
    print(h1, type(h1))
    range_1 = []
    range_2 = []
    with open(f"names-roll.csv", "r") as f:
        reader = csv.reader(f, delimiter=',')
        l = []
        name = []
        for r in reader:
            if r[0] == "Roll":
                continue
            l.append(r[0])
            name.append(r[1])
    for x in range(hem, h1+1):
        jj = 0

        for y in range(0, len(l)):
            if(x >= 10):
                if he + str(x) == l[y]:
                    range_1.append(l[y].upper())
                    jj = 1
                    break
            elif x < 10:
                print(he + "0" + str(x),   l[y])
                if he + "0" + str(x) == l[y].upper():
                    range_1.append(l[y].upper())
                    jj = 1
                    break
        if(jj == 0):
            if x < 10:
                range_2.append(he+"0" + str(x))
            else:
                range_2.append(he + str(x))

  
    print(range_2)
    global l_range
    l_range = range_2
    return
@app.route('/result',methods=['POST', 'GET'])
def result():
    
    name = 'Transcripts uploaded successfully'
   
    generate_transcript()
    return render_template('index.html', name = name)
@app.route('/range',methods=['POST', 'GET'])

def range():
    
    name = 'Transcripts uploaded successfully'
    output = request.form.to_dict()
    global range1
    global range2
    range1 = output['range1']
    range2 = output['range2']
    range1 = range1.upper()
    range2 = range2.upper()
    generate_range_transcript()
    global l_range
    flash(l_range)
    flash('0401ME15')
    
    return render_template('index.html', name = name)


if __name__ == '__main__':
     app.run()

#generate_transcript()
