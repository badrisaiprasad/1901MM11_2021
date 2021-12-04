'''
Team members : 
1901MM11 -- BADRI SAI PRASAD
1901MM25 -- PCS ASWIN KUMAR
'''



from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border , Side
import os
import csv
from flask_mail import Mail, Message
import email, smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import Flask, render_template, request, redirect, url_for
positive = 5
negative = -1
value = 0
location = os.getcwd()
a = []
def send_mail() :
    global location
    os.chdir(location)
    with open(r"sample_input\responses.csv", 'r') as file:
     rows = csv.reader(file)
     addresses = {row[6]:[row[1],row[4]] for row in rows if row[6]!="ANSWER" and row[6]!="Roll Number"}
    for key,values in addresses.items():
     fromaddr = "awesomeaswin138@gmail.com"
     toaddr = values
    
    # instance of MIMEMultipart
     msg = MIMEMultipart()
    
    # storing the senders email address  
     msg['From'] = fromaddr
  
    # storing the receivers email address 
     msg['To'] = ", ".join(values)
    
    # storing the subject 
     msg['Subject'] = "MARKSHEET"
    
    # string to store the body of the mail
     body = " FIND THE MARKSHEET"
    
    # attach the body with the msg instance
     msg.attach(MIMEText(body, 'plain'))
    
    # open the file to be sent 
     print(f"{key}.xlsx")
     filename = f"{key}.xlsx"
     path = os.getcwd() + '\marksheet' + '\\' +filename 
     attachment = open(path, "rb")
    
    # instance of MIMEBase and named as p
     p = MIMEBase('application', 'octet-stream')
    
    # To change the payload into encoded form
     p.set_payload((attachment).read())
    
    # encode into base64
     encoders.encode_base64(p)
    
     p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    
    # attach the instance 'p' to instance 'msg'
     msg.attach(p)
    
    # creates SMTP session
     s = smtplib.SMTP('smtp.gmail.com', 587)
    
    # start TLS for security
     s.starttls()
    
    # Authentication
     s.login(fromaddr, "2002@Aswin")
    
    # Converts the Multipart msg into a string
     text = msg.as_string()
    
    # sending the mail
     s.sendmail(fromaddr, toaddr, text)
    
    # terminating the session
     s.quit()
    return 

def consize() :
    global location
    
    os.chdir(location)
    path = os.getcwd()
    with open('master_roll.csv' , 'r') as f:
     names = list(csv.reader(f))
    with open('responses.csv' , 'r') as g :
     responses  = list(csv.reader(g))
    wb1 = Workbook()
    wb1.create_sheet(index = 0 ,title  = 'concize_marksheet')
    cmsheet = wb1['concize_marksheet']
    for roll in names[1:]:
        os.chdir(path)
        os.chdir('./marksheet')
        lsit = os.listdir()
        mark = openpyxl.load_workbook(roll[0] + '.xlsx')
        quiz = mark['quiz']
        cm = []
        string = ''
        for i in range(7,37) :
            cmsheet.cell(row = 1 , column = i ).value = 'unnamed:'
            pass
        cmsheet['A1'],cmsheet['B1'],cmsheet['C1'],cmsheet['D1'],cmsheet['E1'],cmsheet['F1'],cmsheet['G1'],cmsheet['H1'],cmsheet['AK1']= 'TimeStamp' , 'Email' , 'google_score' , 'Name' , 'IItp_webmail' , 'PHone' , 'Score_after' , 'Roll',' StatusAns'
        for response in responses :
            if(response[3] == roll[1]) :
              cm.append(response[0])
              cm.append(response[1])
              cm.append(response[2])
              cm.append(response[3])
              cm.append(response[4])
              cm.append(response[5])
        global value
        value = quiz.cell(row = 12 , column = 5).value
        cm.append(value)
        for response in responses :
          if(response[3] == roll[1]) :
            cm.append(response[6])
            for i in range(7,35) :
             cm.append(response[i])
        stri = '[' + str(quiz.cell(row = 10 , column = 2).value) + ',' + str(quiz.cell(row = 10 , column = 3).value) + ',' + str(quiz.cell(row = 10 , column = 4).value) + ']'
        cm.append(stri)
        cmsheet.append(cm)
    cmsheet.delete_rows(2)   
    wb1.save('concize_marksheet.xlsx')


    return 


def generate_marksheet() :
 font1 = Font(color='000000FF' , size= '12' , name= 'Century')       #blue for  correct options 
 font2 = Font(color='FF0000' , size= '12' , name= 'Century')         #red colour for incorrect answer
 font3 = Font(color='008000' , size='12' ,name= 'Century')         #green colour for correct answer
 os.mkdir('./marksheet')
 path = os.getcwd()
#os.chdir('./sample_input')
 with open('master_roll.csv' , 'r') as f:
  names = list(csv.reader(f))
  with open('responses.csv' , 'r') as g :
     responses  = list(csv.reader(g))
 wb1 = Workbook()
 wb1.create_sheet(index = 0 ,title  = 'concize_marksheet')

 cmsheet = wb1['concize_marksheet']
 for roll in names[1:]:
    os.chdir(path)
    os.chdir('./marksheet')
    wb= Workbook()
    wb.create_sheet(index = 0 ,title  = 'quiz')
    cm = []

    cmsheet['A1'],cmsheet['B1'],cmsheet['C1'],cmsheet['D1'],cmsheet['E1'],cmsheet['F1'],cmsheet['G1'],cmsheet['H1'] = 'TimeStamp' , 'Email' , 'google_score' , 'Name' , 'IItp_webmail' , 'PHone' , 'Score_after' , 'Roll'
    correct_answers = 0
    incorrect_answers = 0
    not_attempted = 0 
    quiz = wb['quiz']
    quiz['A6'] ,quiz['A7'] , quiz['D6'] , quiz['E6'] = 'Name:' , 'Roll Number:' , 'Exam:' , 'quiz'
    quiz['A10'] ,quiz['A11'] , quiz['A12'] = 'No.' , 'Marking:' , 'Total'
    quiz['A10'].font ,quiz['A11'].font , quiz['A12'].font = Font(bold = True,size='12' ,name= 'Century'),Font(bold = True,size='12' ,name= 'Century'),Font(bold = True,size='12' , name= 'Century')
    quiz['A15'] , quiz['B15'] = 'Student Ans' , 'Correct Ans'
    quiz['D15'] , quiz['E15'] = 'Student Ans' , 'Correct Ans'
    quiz['A15'].font , quiz['B15'].font =  Font(bold = True , size='12' , name= 'Century'),Font(bold = True ,size='12' , name= 'Century')
    quiz['D15'].font , quiz['E15'].font =  Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century')
    global positive
    global negative
    quiz['B11'],quiz['C11'] , quiz['D11']= positive, negative , 0
    quiz['B11'].font,quiz['C11'].font  = font3 , font2
    quiz['B9'],quiz['C9'],quiz['D9'],quiz['E9'] = 'Right' , 'Wrong' , 'Not Attempt' , 'Max'
    quiz['B9'].font,quiz['C9'].font,quiz['D9'].font,quiz['E9'].font = Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century'),Font(bold = True,size='12' , name= 'Century') , Font(bold = True,size='12' , name= 'Century')
    quiz['A5'] = 'Mark Sheet'
    quiz['A5'].font = Font( underline= 'single' , size='18')
    quiz['B16'] , quiz['B17'] , quiz['B18'],quiz['B19']  = 'Option A' , 'Option D' , 'Option B' , 'Option C'
    quiz['B20'] , quiz['B21'],quiz['B22'] , quiz['B23']  = 'Option B' , 'Option C' , 'Option D' , 'Option D'
    quiz['B24'] , quiz['B25'] , quiz['B26'],quiz['B27']  = 'Option A' , 'Option A' , 'Option C' , 'Option A'
    quiz['B28'] , quiz['B29'],quiz['B30'] , quiz['B31']  = 'Option D' , 'Option D' , 'Option B' , 'Option D'
    quiz['B32'] , quiz['B33'] , quiz['B34'],quiz['B35']  = 'Option C' , 'Option D' , 'Option B' , 'Option D'
    quiz['B36'] , quiz['B37'],quiz['B38'] , quiz['B39'] , quiz['B40']  = 'Option A' , 'Option A' , 'Option A' , 'Option D' , 'Option D'
    quiz['E16'] , quiz['E17'] , quiz['E18']  = 'Option A' , 'Option C' , 'Option D' 
    quiz['E16'].font , quiz['E17'].font , quiz['E18'].font  = font1 , font1 , font1
    for i in range(16,41) :
        string = 'B' + str(i)
        quiz[string].font = font1 
    for response in responses :
        if(response[3] == roll[1]) :
              quiz['A16'] , quiz['A17'] , quiz['A18'],quiz['A19']  = response[7] , response[8] , response[9] ,response[10]
              quiz['A20'] , quiz['A21'],quiz['A22'] , quiz['A23']  = response[11] , response[12] , response[13] ,response[14]
              quiz['A24'] , quiz['A25'] , quiz['A26'],quiz['A27']  = response[15] , response[16] , response[17] ,response[18]
              quiz['A28'] , quiz['A29'],quiz['A30'] , quiz['A31']  = response[19] , response[20] , response[21] ,response[22]
              quiz['A32'] , quiz['A33'] , quiz['A34'],quiz['A35']  = response[23] , response[24] , response[25] ,response[26]
              quiz['A36'] , quiz['A37'],quiz['A38'] , quiz['A39'] , quiz['A40']  = response[27] , response[28] , response[29] ,response[30],response[31]
              quiz['D16'] , quiz['D17'] , quiz['D18']  = response[32] , response[33] , response[34] 
              semail = response[1]
              cm.append(response[0])
              cm.append(response[1])
              cm.append(response[2])
              cm.append(response[3])
              cm.append(response[4])
              cm.append(response[5])
    for i in range(16,41) :
        string1 = 'A' + str(i)
        string2 = 'B' + str(i)
        if(quiz[string1].value == quiz[string2].value) :
            quiz[string1].font = font3
            correct_answers = correct_answers + 1
        elif (quiz[string1].value != quiz[string2].value) and quiz[string1].value != '' :   
            quiz[string1].font = font2     
            incorrect_answers = incorrect_answers + 1
        else :
            not_attempted = not_attempted + 1
    for i in range(16,19) :
        string1 = 'D' + str(i)
        string2 = 'E' + str(i) 
        if(quiz[string1].value == quiz[string2].value) :
            quiz[string1].font = font3
            correct_answers = correct_answers + 1
        elif (quiz[string1].value != quiz[string2].value) and quiz[string1].value != '' :   
            quiz[string1].font = font2     
            incorrect_answers = incorrect_answers + 1
        else :
            not_attempted = not_attempted + 1        
    quiz['B10'] ,quiz['C10'] , quiz['D10'] ,quiz['E10'] = correct_answers , incorrect_answers , not_attempted , 28
    global a
    a.append(correct_answers)
    a.append(incorrect_answers)
    a.append(incorrect_answers)
    quiz['B12'] ,quiz['C12']  = (quiz['B10'].value * quiz['B11'].value),(quiz['C10'].value * quiz['C11'].value)
    quiz['E12'] = quiz['B12'].value + quiz['C12'].value
    quiz['E12'] = str(quiz['E12'].value) + '/140'
    quiz['E12'].font = font1
    quiz['B12'].font , quiz['B10'].font = font3 , font3
    quiz['C12'].font , quiz['C10'].font = font2 , font2
    quiz['B7'] , quiz['B6'] = roll[0] , roll[1]
    quiz['B6'].font,quiz['B7'].font,quiz['E6'].font = Font(bold = True , size='12' , name= 'Century'),Font(bold = True , size='12' , name= 'Century'),Font(bold = True , size='12' , name= 'Century') 
    quiz.column_dimensions['A'].width = 16.89
    quiz.column_dimensions['B'].width = 16.89
    quiz.column_dimensions['C'].width = 16.89
    quiz.column_dimensions['D'].width = 16.89
    quiz.column_dimensions['E'].width = 16.89
    quiz.merge_cells('A5:E5')
    quiz.merge_cells('B6:C6')
    thin_border = Border(left = Side(style='thin') ,right = Side(style='thin') ,top = Side(style='thin') ,bottom = Side(style='thin')  )
    for i in range(15,41) :
     quiz.cell(row = i , column = 1).border = thin_border
     quiz.cell(row = i , column = 2).border = thin_border
    for i in range(9,13) :
     quiz.cell(row = i , column = 1).border = thin_border
     quiz.cell(row = i , column = 2).border = thin_border
     quiz.cell(row = i , column = 3).border = thin_border
     quiz.cell(row = i , column = 4).border = thin_border
     quiz.cell(row = i , column = 5).border = thin_border
    for i in range(15,19) :
     quiz.cell(row = i , column = 4).border = thin_border
     quiz.cell(row = i , column = 5).border = thin_border
    #quiz.column_dimensions['B'].width = 25  
    for i in range(quiz.max_column) :
     for row in quiz[2:quiz.max_row]:
       cell = row[i]
       cell.alignment = Alignment(horizontal='center') 
    img = openpyxl.drawing.image.Image( path + '/iitp logo.png')
    img.anchor = 'A1'
    img.width = 600
    img.height = 85
    quiz.add_image(img)
    global value
    value = quiz['E12'].value
    cm.append(quiz['E12'].value)
    for response in responses :
        if(response[3] == roll[1]) :
          cm.append(response[6])
          for i in range(7,35) :
             cm.append(response[i])
    cmsheet.append(cm)
    wb.save(roll[0] + '.xlsx')
    
  # wb1.save('concize_marksheet.xlsx')
    '''
    fromadd= 'awesomeaswin138@gmail.com'
    toadd = semail
    msg = MIMEMultipart()
    msg["From"] = 'awesomeaswin138@gmail.com'
    msg["To"] = semail
    msg["Subject"] = 'test email'
    body = 'assignment'
    msg.attach(MIMEText(body, "plain"))
    filepath = os.getcwd()
    filename = roll[0] + '.xlsx'
    file = open(filepath + '\\' + filename , 'rb')

    payload = MIMEBase('application' , 'octet-stream')
    payload.set_payload(file.read())
    file.close()
    encoders.encode_base64(payload)
    payload.add_header('Content-Disposition' , 'attachment' , filename = filename)
    msg.attach(payload)

    s = smtplib.SMTP('smtp@gmail.com' , 587)
    s.starttls()
    s.login('awesomeaswin138@gmail.com','2002@Aswin')
    text = msg.as_string()
    s.sendmail('awesomeaswin138@gmail.com' , semail , text)
    s.quit()
    
    '''
    '''
    msg = MIMEMultipart()
    msg["From"] = 'awesomeaswin138@gmail.com'
    msg["To"] = semail
    msg["Subject"] = 'test email'
    msg["Bcc"] = semail  # Recommended for mass emails
    body = ''' '''
    msg.attach(MIMEText(body, "plain"))
    filepath = os.getcwd()
    filename = roll[0] + '.xlsx'
    file = open(filepath + '\\' + filename , 'rb')

    payload = MIMEBase('application' , 'octet-stream')
    payload.set_payload(file.read())
    file.close()
    encoders.encode_base64(payload)
    payload.add_header('Content-Disposition' , 'attachment' , filename = filename)
    msg.attach(payload)

    server = gmail_login(username = 'awesomeaswin138@gmail.com' , password = '2002@Aswin')
    server.send_message(msg)
    server.quit()
    print('email has sent successfully')
    '''
app = Flask(__name__)
mail=Mail(app)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/', methods=['POST'])
def upload_file():
    uploaded_file = request.files['file']
    if uploaded_file.filename != '':
        uploaded_file.save(uploaded_file.filename)
    return redirect(url_for('index'))
@app.route('/result',methods=['POST', 'GET'])
def result():
    marksheet = request.form.to_dict()
    
    name = 'File uploaded successfully'
    
    global positive
    global negative
    positive = int(marksheet['positive marks'])
    negative = int(marksheet['negative marks'])
    
    generate_marksheet()
    return render_template('index.html', name = name)
@app.route('/sendMail',methods=['POST', 'GET'])
def sendMail():
    
    
    
    send_mail()
    return render_template('index.html')
@app.route('/marksheet',methods=['POST', 'GET'])
def marksheet():
    
    
    
    consize()
    return render_template('index.html')


if __name__ == '__main__':
   app.run()




#generate_marksheet()       